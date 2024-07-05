# -*- coding: utf-8 -*-

import base64
import os
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile

from odoo import api, fields, models
from openpyxl import Workbook, load_workbook

# Porcentajes para las barras de cargas
AVAILABLE_PRIORITIES = [
    ('na', 'N/A - No aplica'),
    ('0_porcent', '0% - No documentado / No existente'),
    ('25_porcent', '25% - Aplicado / No documentado'),
    ('50_porcent', '50% - Documentado / No aplicado'),
    ('75_porcent', '75% - Aplicado y documentado'),
    ('100_porcent', '100% - Aplicado, documentado y controlado')]

FIELDS = ('diagnostic4_ids', 'diagnostic4_1_ids', 'diagnostic4_2_ids', 'diagnostic4_3_1_ids', 'diagnostic4_3_2_ids',
          'diagnostic5_1_ids', 'diagnostic5_2_ids', 'diagnostic5_1_1_ids', 'diagnostic5_1_2_ids', 'diagnostic5_2_1_ids',
          'diagnostic6_1_ids', 'diagnostic6_2_ids', 'diagnostic6_1_1_ids', 'diagnostic6_1_2_ids', 'diagnostic6_2_1_ids',
          'diagnostic7_1_1_ids', 'diagnostic7_1_2_ids', 'diagnostic7_1_3_ids', 'diagnostic7_1_4_ids', 'diagnostic7_1_5_1_ids', 'diagnostic7_1_5_2_ids', 'diagnostic7_1_6_ids', 'diagnostic7_2_1_ids', 'diagnostic7_3_1_ids', 'diagnostic7_4_1_ids', 'diagnostic7_5_1_ids', 'diagnostic7_5_2_ids', 'diagnostic7_5_3_1_ids', 'diagnostic7_5_3_2_ids',
          'diagnostic8_1_1_ids', 'diagnostic8_2_1_ids', 'diagnostic8_2_2_ids', 'diagnostic8_2_2_3_1_ids', 'diagnostic8_2_2_3_2_ids', 'diagnostic8_2_4_ids', 'diagnostic8_3_1_ids', 'diagnostic8_3_2_ids', 'diagnostic8_3_3_ids', 'diagnostic8_3_4_ids', 'diagnostic8_3_5_ids', 'diagnostic8_3_6_ids', 'diagnostic8_4_1_ids', 'diagnostic8_4_2_ids', 'diagnostic8_4_3_ids', 'diagnostic8_5_1_ids', 'diagnostic8_5_2_ids', 'diagnostic8_5_3_ids', 'diagnostic8_5_4_ids', 'diagnostic8_5_5_ids', 'diagnostic8_5_6_ids', 'diagnostic8_6_1_ids', 'diagnostic8_7_1_ids', 'diagnostic8_7_2_ids',
          'diagnostic9_1_1_ids', 'diagnostic9_1_2_ids', 'diagnostic9_1_3_ids', 'diagnostic9_2_1_ids', 'diagnostic9_2_2_ids', 'diagnostic9_3_1_ids', 'diagnostic9_3_2_ids', 'diagnostic9_3_3_ids',
          'diagnostic10_1_1_ids', 'diagnostic10_2_1_ids', 'diagnostic10_2_2_ids', 'diagnostic10_3_1_ids')


class XLSHelper(models.Model):
    """Modelo para guardar la documentación de Soy Calidad en formato xlsx
    Esto para evitar cambios en las reglas de acceso
    """
    _name = 'hola_calidad.xls_helper'
    _description = 'Ayudante binario para archivos xlsx'

    name = fields.Char('Nombre')
    datas = fields.Binary('File', readonly=True)
    datas_fname = fields.Char('Filename', readonly=True)
    date_validate = fields.Datetime(string=u'Fecha evaluación')


class Clause(models.Model):
    _name = 'hola_calidad.clause'
    _description = "Claúsulas"
    # _order = "chapter asc"

    question = fields.Text(string=u'Pregunta ref.', required=True)
    name = fields.Char(string=u'Nombre', required=True)
    complete_name = fields.Text(string=u'Descripción', required=True)
    chapter = fields.Selection(
        string=u'Capítulo',
        selection=[
            ('4_context', 'Contexto de la organización'),
            ('5_leadership', 'Liderazgo'),
            ('6_planning', 'Planificación'),
            ('7_support', 'Apoyo'),
            ('8_operation', 'Operación'),
            ('9_evaluation', 'Evaluación del desempeño'),
            ('10_improvement', 'Mejora')],
        required=True,
    )


class Requirement(models.Model):
    _name = 'hola_calidad.requirement'
    _description = "Requirimientos"

    name = fields.Char(string=u'Nombre', required=True)
    complete_name = fields.Text(string=u'Descripción', required=True)
    info = fields.Text(string=u'Interpretación', store=True)
    clause_id = fields.Many2one(
        string=u'Clausula', comodel_name='hola_calidad.clause', required=True)
    chapter = fields.Selection(
        string=u'Capítulo', related='clause_id.chapter', store=True)
    position_excel = fields.Char(string=u'Posición en excel')


class DiagnosticLine(models.Model):
    _name = 'hola_calidad.diagnostic.line'
    _description = 'Linea de análisis de calidad'

    requirement_name = fields.Char(related='requirement_id.name')
    # FIX
    diagnostic4_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic4_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic4_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic4_3_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic4_3_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')

    diagnostic5_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic5_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic5_1_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic5_1_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic5_2_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')

    diagnostic6_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic6_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic6_1_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic6_1_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic6_2_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')

    diagnostic7_1_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_1_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_1_3_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_1_4_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_1_5_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_1_5_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_1_6_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_2_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_3_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_4_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_5_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_5_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_5_3_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic7_5_3_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')

    diagnostic8_1_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_2_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_2_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_2_2_3_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_2_2_3_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_2_4_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_3_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_3_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_3_3_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_3_4_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_3_5_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_3_6_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_4_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_4_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_4_3_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_5_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_5_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_5_3_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_5_4_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_5_5_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_5_6_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_6_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_7_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic8_7_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')

    diagnostic9_1_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic9_1_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic9_1_3_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic9_2_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic9_2_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic9_3_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic9_3_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic9_3_3_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')

    diagnostic10_1_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic10_2_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic10_2_2_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')
    diagnostic10_3_1_id = fields.Many2one(
        string=u'Diagnostico', comodel_name='hola_calidad.diagnostic', ondelete='cascade')

    requirement_id = fields.Many2one(
        string=u'Requisito', comodel_name='hola_calidad.requirement',)
    clause_id = fields.Text(string=u'Clausula ID', store=True)
    clause = fields.Many2one(
        string=u'Claúsulas', comodel_name='hola_calidad.clause', ondelete='cascade')

    info = fields.Text(string=u'Interpretación',
                       help="here is my message", store=True)

    name = fields.Text(string=u'Nombre requirimiento', store=True)
    # qualification = fields.Selection(
    #     string=u'Calificación',
    #     required=True,
    #     selection=[('0_porcent', '0%'), ('25_porcent', '25%'), ('50_porcent', '50%'), ('75_porcent', '75%'), ('100_porcent', '100%'), ('na', 'N/A')],
    #     default='na'
    # )
    qualification = fields.Selection(AVAILABLE_PRIORITIES,
                                     index=True,
                                     string=u'Calificación',
                                     required=True,
                                     default='na')

    observation = fields.Text(string=u'Observaciones')
    is_page = fields.Boolean('Is a page?')
    display_type = fields.Selection([
        ('line_section', 'Section'),
        ('line_note', 'Note'),
    ], default=False, help="Technical field for UX purpose.")

    @api.model
    def create(self, values):
        if not values.get('name'):
            vcomplete_name = self.env['hola_calidad.requirement'].browse(
                values.get('requirement_id')).complete_name
            values['name'] = vcomplete_name
        result = super(DiagnosticLine, self).create(values)
        return result

    @api.onchange('requirement_id')
    def _onchange_requirement_id(self):
        self.name = self.requirement_id.complete_name

    # def show_info_hover(self):
    #     print ("Mostrar INFO")
    #     print (self)

        # raise (_('Header!'), _('Message.'))


class Diagnostic(models.Model):
    _name = 'hola_calidad.diagnostic'
    _description = "Diagnostico"

    name = fields.Char(string=u'Nombre', required=True,
                       default=lambda self: "Análisis de calidad")
    user_id = fields.Many2one(
        string='Responsable',
        comodel_name='res.users',
        ondelete='cascade',
        default=lambda self: self.env.user and self.env.user.id or False,
    )

    company_id = fields.Many2one(string=u'Compañia', comodel_name='res.company', required=True,
                                 domain=lambda self: [('id', 'in', self.env.user.company_ids.ids)], default=lambda self: self.env.user.company_id.id)
    date_diagnostic = fields.Datetime(
        string=u'Fecha creación', default=fields.Datetime.now, required=True)
    date_validate = fields.Datetime(
        string=u'Fecha evaluación', related='xls_helper.date_validate')

    all_clause = fields.Many2many(
        comodel_name='hola_calidad.clause', string=u'Clausulas')

    diagnostic4_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic4_id',)
    diagnostic4_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic4_1_id',)
    diagnostic4_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic4_2_id',)
    diagnostic4_3_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic4_3_1_id',)
    diagnostic4_3_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic4_3_2_id',)

    diagnostic5_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic5_1_id',)
    diagnostic5_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic5_2_id',)
    diagnostic5_1_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic5_1_1_id',)
    diagnostic5_1_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic5_1_2_id',)
    diagnostic5_2_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic5_2_1_id',)

    diagnostic6_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic6_1_id',)
    diagnostic6_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic6_2_id',)
    diagnostic6_1_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic6_1_1_id',)
    diagnostic6_1_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic6_1_2_id',)
    diagnostic6_2_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic6_2_1_id',)

    diagnostic7_1_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_1_1_id',)
    diagnostic7_1_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_1_2_id',)
    diagnostic7_1_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_1_3_id',)
    diagnostic7_1_4_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_1_4_id',)
    diagnostic7_1_5_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_1_5_1_id',)
    diagnostic7_1_5_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_1_5_2_id',)
    diagnostic7_1_6_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_1_6_id',)

    diagnostic7_2_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_2_1_id',)

    diagnostic7_3_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_3_1_id',)

    diagnostic7_4_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_4_1_id',)

    diagnostic7_5_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_5_1_id',)

    diagnostic7_5_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_5_2_id',)

    diagnostic7_5_3_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_5_3_1_id',)
    diagnostic7_5_3_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic7_5_3_2_id',)

    diagnostic8_1_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_1_1_id', )
    diagnostic8_2_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_2_1_id', )
    diagnostic8_2_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_2_2_id', )
    diagnostic8_2_2_3_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_2_2_3_1_id', )
    diagnostic8_2_2_3_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_2_2_3_2_id', )
    diagnostic8_2_4_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_2_4_id', )
    diagnostic8_3_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_3_1_id', )
    diagnostic8_3_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_3_2_id', )
    diagnostic8_3_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_3_3_id', )
    diagnostic8_3_4_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_3_4_id', )
    diagnostic8_3_5_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_3_5_id', )
    diagnostic8_3_6_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_3_6_id', )
    diagnostic8_4_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_4_1_id', )
    diagnostic8_4_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_4_2_id', )
    diagnostic8_4_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_4_3_id', )
    diagnostic8_5_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_5_1_id', )
    diagnostic8_5_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_5_2_id', )
    diagnostic8_5_3_ids = fields.One2many(
        string=u'Lneas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_5_3_id', )
    diagnostic8_5_4_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_5_4_id', )
    diagnostic8_5_5_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_5_5_id', )
    diagnostic8_5_6_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_5_6_id', )
    diagnostic8_6_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_6_1_id', )
    diagnostic8_7_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_7_1_id', )
    diagnostic8_7_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic8_7_2_id', )

    diagnostic9_1_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_1_1_id', )
    diagnostic9_1_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_1_2_id', )
    diagnostic9_1_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_1_3_id', )
    diagnostic9_2_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_2_1_id', )
    diagnostic9_2_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_2_2_id', )
    diagnostic9_3_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_3_1_id', )
    diagnostic9_3_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_3_2_id', )
    diagnostic9_3_3_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic9_3_3_id', )

    diagnostic10_1_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic10_1_1_id', )
    diagnostic10_2_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic10_2_1_id', )
    diagnostic10_2_2_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic10_2_2_id', )
    diagnostic10_3_1_ids = fields.One2many(
        string=u'Lineas', comodel_name='hola_calidad.diagnostic.line', inverse_name='diagnostic10_3_1_id', )

    state = fields.Selection(
        string=u'Estado',
        selection=[('draft', 'Previo'),
                   ('evaluate', 'Detallado'),
                   ('validate', 'Culminado')],
        default='draft',
    )

    def get_diagnostic_values(self, diagnostic_list):
        total_100 = total_75 = total_50 = total_25 = total_0 = total_na = 0
        for line in diagnostic_list:
            if line.qualification == 'na':
                total_na += 1
            if line.qualification == '100_porcent':
                total_100 += 1
            if line.qualification == '75_porcent':
                total_75 += 1
            if line.qualification == '50_porcent':
                total_50 += 1
            if line.qualification == '25_porcent':
                total_25 += 1
            if line.qualification == '0_porcent':
                total_0 += 1
        return [total_na, total_100, total_75, total_50, total_25, total_0]

    diagnostic4_ids_100 = fields.Integer(
        string=u'Total Contexto 100%', compute='_get_diagnostic')
    diagnostic4_ids_75 = fields.Integer(
        string=u'Total Contexto 75%', compute='_get_diagnostic')
    diagnostic4_ids_50 = fields.Integer(
        string=u'Total Contexto 50%', compute='_get_diagnostic')
    diagnostic4_ids_25 = fields.Integer(
        string=u'Total Contexto 25%', compute='_get_diagnostic')
    diagnostic4_ids_0 = fields.Integer(
        string=u'Total Contexto 0%', compute='_get_diagnostic')

    diagnostic5_ids_100 = fields.Integer(
        string=u'Total Liderazgo 100%', compute='_get_diagnostic')
    diagnostic5_ids_75 = fields.Integer(
        string=u'Total Liderazgo 75%', compute='_get_diagnostic')
    diagnostic5_ids_50 = fields.Integer(
        string=u'Total Liderazgo 50%', compute='_get_diagnostic')
    diagnostic5_ids_25 = fields.Integer(
        string=u'Total Liderazgo 25%', compute='_get_diagnostic')
    diagnostic5_ids_0 = fields.Integer(
        string=u'Total Liderazgo 0%', compute='_get_diagnostic')

    diagnostic6_ids_100 = fields.Integer(
        string=u'Total Planificacion 100%', compute='_get_diagnostic')
    diagnostic6_ids_75 = fields.Integer(
        string=u'Total Planificacion 75%', compute='_get_diagnostic')
    diagnostic6_ids_50 = fields.Integer(
        string=u'Total Planificacion 50%', compute='_get_diagnostic')
    diagnostic6_ids_25 = fields.Integer(
        string=u'Total Planificacion 25%', compute='_get_diagnostic')
    diagnostic6_ids_0 = fields.Integer(
        string=u'Total Planificacion 0%', compute='_get_diagnostic')

    diagnostic7_ids_100 = fields.Integer(
        string=u'Total Apoyo 100%', compute='_get_diagnostic')
    diagnostic7_ids_75 = fields.Integer(
        string=u'Total Apoyo 75%', compute='_get_diagnostic')
    diagnostic7_ids_50 = fields.Integer(
        string=u'Total Apoyo 50%', compute='_get_diagnostic')
    diagnostic7_ids_25 = fields.Integer(
        string=u'Total Apoyo 25%', compute='_get_diagnostic')
    diagnostic7_ids_0 = fields.Integer(
        string=u'Total Apoyo 0%', compute='_get_diagnostic')

    diagnostic8_ids_100 = fields.Integer(
        string=u'Total Operación 100%', compute='_get_diagnostic')
    diagnostic8_ids_75 = fields.Integer(
        string=u'Total Operación 75%', compute='_get_diagnostic')
    diagnostic8_ids_50 = fields.Integer(
        string=u'Total Operación 50%', compute='_get_diagnostic')
    diagnostic8_ids_25 = fields.Integer(
        string=u'Total Operación 25%', compute='_get_diagnostic')
    diagnostic8_ids_0 = fields.Integer(
        string=u'Total Operación 0%', compute='_get_diagnostic')

    diagnostic9_ids_100 = fields.Integer(
        string=u'Total Desempeño 100%', compute='_get_diagnostic')
    diagnostic9_ids_75 = fields.Integer(
        string=u'Total Desempeño 75%', compute='_get_diagnostic')
    diagnostic9_ids_50 = fields.Integer(
        string=u'Total Desempeño 50%', compute='_get_diagnostic')
    diagnostic9_ids_25 = fields.Integer(
        string=u'Total Desempeño 25%', compute='_get_diagnostic')
    diagnostic9_ids_0 = fields.Integer(
        string=u'Total Desempeño 0%', compute='_get_diagnostic')

    diagnostic10_ids_100 = fields.Integer(
        string=u'Total Mejora 100%', compute='_get_diagnostic')
    diagnostic10_ids_75 = fields.Integer(
        string=u'Total Mejora 75%', compute='_get_diagnostic')
    diagnostic10_ids_50 = fields.Integer(
        string=u'Total Mejora 50%', compute='_get_diagnostic')
    diagnostic10_ids_25 = fields.Integer(
        string=u'Total Mejora 25%', compute='_get_diagnostic')
    diagnostic10_ids_0 = fields.Integer(
        string=u'Total Mejora 0%', compute='_get_diagnostic')

    diagnostic4_ids_total = fields.Integer(
        string=u'Total Punto 4', compute='_get_diagnostic', store=True)
    diagnostic5_ids_total = fields.Integer(
        string=u'Total Punto 5', compute='_get_diagnostic', store=True)
    diagnostic6_ids_total = fields.Integer(
        string=u'Total Punto 6', compute='_get_diagnostic', store=True)
    diagnostic7_ids_total = fields.Integer(
        string=u'Total Punto 7', compute='_get_diagnostic', store=True)
    diagnostic8_ids_total = fields.Integer(
        string=u'Total Punto 8', compute='_get_diagnostic', store=True)
    diagnostic9_ids_total = fields.Integer(
        string=u'Total Punto 9', compute='_get_diagnostic', store=True)
    diagnostic10_ids_total = fields.Integer(
        string=u'Total Punto 10', compute='_get_diagnostic', store=True)

    def get_diagnostic_values(self, diagnostic_list):
        total_100 = total_75 = total_50 = total_25 = total_0 = total_na = 0

        for line in diagnostic_list:
            line_t = getattr(self, line)
            for line_ in line_t:
                if line_.qualification == 'na':
                    total_na += 1
                if line_.qualification == '100_porcent':
                    total_100 += 1
                if line_.qualification == '75_porcent':
                    total_75 += 1
                if line_.qualification == '50_porcent':
                    total_50 += 1
                if line_.qualification == '25_porcent':
                    total_25 += 1
                if line_.qualification == '0_porcent':
                    total_0 += 1
        total_in_partials = [total_na, total_100,
                             total_75, total_50, total_25, total_0]
        sum_total = sum(total_in_partials)
        return [total_na, total_100, total_75, total_50, total_25, total_0, sum_total]

    def _get_diagnostic(self):
        fields_ = dir(self)
        field_suffixes = ['4', '5', '6', '7', '8', '9', '10']
        field_string = 'diagnostic{}_ids_{}'
        field_string_ = 'diagnostic{}'
        field_total = 'diagnostic{}_ids_total'

        for suffix in field_suffixes:
            diagnostic_list = filter(lambda x: x.startswith(
                field_string_.format(suffix)), FIELDS)
            res = self.get_diagnostic_values(diagnostic_list)
            setattr(self, field_string.format(suffix, '100'), res[1])
            setattr(self, field_string.format(suffix, '75'), res[2])
            setattr(self, field_string.format(suffix, '50'), res[3])
            setattr(self, field_string.format(suffix, '25'), res[4])
            setattr(self, field_string.format(suffix, '0'), res[5])
            setattr(self, field_total.format(suffix), res[6])

    def _default_diagnostic_line_ids(self, vchapter):
        """ 
        """
        requirements = self.env['hola_calidad.requirement'].search(
            [('chapter', '=', vchapter)])

        lines = [(5, 0, 0)]
        for req in requirements:
            if req.clause_id.id == '4.1':
                data = {
                    'name': req.complete_name,
                    'requirement_id': req.id,
                    'clause_id': req.clause_id.id,
                    'qualification': 'na',
                }
                lines.append((0, 0, data))
        return lines

    def _default_diagnostic_line_ids_v2(self, vchapter, ids_clause):
        """ Devuelve los requisitos de las clausulas, de acuerdo
        a su capitulo y id_req
        """
        requirements = self.env['hola_calidad.requirement'].search(
            [('name', '=like', vchapter + '%')])

        lines = [(5, 0, 0)]
        print([x.name for x in requirements])
        for req in requirements:
            if (req.clause_id.id in ids_clause):
                data = {
                    'info': req.info,
                    'name': req.complete_name,
                    'requirement_id': req.id,
                    'clause_id': req.clause_id.id,
                    'qualification': 'na',
                }
                lines.append((0, 0, data))
        return lines

    def _get_xls_helper(self):
        res = self.env['hola_calidad.xls_helper'].search([])
        return res

    datas = fields.Binary('File', readonly=True)
    datas_fname = fields.Char('Filename', readonly=True)
    xls_helper = fields.Many2one(comodel_name='hola_calidad.xls_helper',
                                 string='Soy Calidad', readonly=True, default=_get_xls_helper)

    def evaluate_diagnostic(self):
        paths = os.path.realpath(__file__)
        dirname = os.path.dirname(os.path.dirname(paths))
        newdir = os.path.join(dirname, 'data')
        workbook = load_workbook(newdir+'/data.xlsx')

        # user = self.env['res.users'].browse(self.env.uid)
        user = self.env.user
        time = datetime.now()
        #data = self.env['hola_calidad.diagnostic'].browse(self._context.get('active_ids',[]))
        self.xls_helper.write({'date_validate': datetime.now()})
        if time:
            filename = self.name + ' ' + \
                str(time.strftime("%Y-%m-%d %H:%M %p"))
        else:
            filename = self.name

        sheets = workbook.sheetnames
        sheet = workbook[sheets[1]]
        cont = 15
        cont_relle = 0

        for diagnostics in [getattr(self, x) for x in FIELDS]:
            for diagnostic_line in diagnostics:
                p_excel = diagnostic_line.requirement_id.position_excel
                print("p_excel------>", p_excel)
                if p_excel and cont < 416:
                    number = p_excel[1:]
                    if number != str(cont):
                        # for i=number; in diagnostics: cell1 = sheet['B'+str(cont)]
                        # if cont+1 in [17, 18,22,23] 31
                        i = cont
                        tmp = 1
                        tmpfinal = True
                        tmp1 = 1
                        while tmpfinal == True and cont < 416 and i < 416:
                            if i not in [17, 18, 22, 23, 29, 30, 41, 44, 45, 46, 58, 63, 64, 70, 74, 75, 82, 83, 84, 90, 94, 95, 98, 106, 112, 113, 118, 119, 120, 124, 126, 128, 130, 131, 136, 141, 145, 146, 151, 152, 157, 158, 164, 165, 168, 169, 173, 174, 177, 184, 185, 186, 198, 199, 205, 209, 210, 220, 223, 225, 226, 228, 239, 248, 255, 261, 267, 268, 275, 282, 292, 293, 304, 308, 312, 314, 320, 323, 329, 330, 338, 343, 344, 345, 352, 355, 364, 365, 370, 377, 378, 380, 387, 392, 393, 398, 399, 410, 413] and i != int(number):
                                if cont < 416 and i < 416:
                                    # comvertir a string
                                    cell1 = sheet['B'+str(i)]
                                    cell1.value = 'X'
                                    tmp = 0
                                    i = i+1
                                    tmp1 = 0
                            else:
                                if tmp == 0:
                                    print("vista tmp y i------>", i)
                                    tmpfinal = False
                                    i = int(number)
                                    break

                            if number == str(i):  # 31
                                tmpfinal = False
                                i = int(number)
                                break
                            else:
                                if tmp1 != 0:
                                    i = i+1
                        cont = i
                    if number == str(cont):
                        number = p_excel[1:]
                        print("number------>", number)
                        if diagnostic_line.qualification == 'na':
                            letter = 'G'
                        elif diagnostic_line.qualification == '0_porcent':
                            letter = 'B'
                        elif diagnostic_line.qualification == '25_porcent':
                            letter = 'C'
                        elif diagnostic_line.qualification == '50_porcent':
                            letter = 'D'
                        elif diagnostic_line.qualification == '75_porcent':
                            letter = 'E'
                        elif diagnostic_line.qualification == '100_porcent':
                            letter = 'F'

                        cell = sheet[letter+number]
                        print(cell)
                        cell.value = 'X'
                        # cont=int(number) #comvertir a entero
                        cont = cont+1
                    if diagnostic_line.observation:
                        cell2 = sheet['H'+number]
                        cell2.value = diagnostic_line.observation

        workbook.close()
        
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                xls_filelike = BytesIO(f.read())

        out = base64.encodestring(xls_filelike.getvalue())

        self.xls_helper.write({'datas': out, 'datas_fname': filename})
        filename += '%2Exlsx'

        return {
            'type': 'ir.actions.act_url',
            'target': 'new',
            'url': 'web/content/?model='+self.xls_helper._name+'&id='+str(self.xls_helper.id)+'&field=datas&download=true&filename='+filename,
        }


    def search_all_clauseId_in_actualRequierments(self):
        fields = FIELDS
        """ Los padres de los requisitos actuales"""
        lista_clausulas = []
        for field in fields:
            for line in getattr(self, field):
                clau_id = line.clause_id
                if clau_id not in lista_clausulas:
                    lista_clausulas.append(clau_id)

        return lista_clausulas

    # ids_clause_list_old_global = all_clause.ids
    def change_state_eval(self):
        """ Aqui se verfica cada vez que se añade una clasula,
        para traer los requisitos relacionadas a esa clausula.
        Cambia el estado de draft a evalate,  cuando son seleccionadas las clausulas
        """

        ids_clause_list_all = self.all_clause.ids

        clausulas_analis_prev = self.search_all_clauseId_in_actualRequierments()
        if (len(ids_clause_list_all)):
            ids_clause_list = []

            print("LISTA DE Existentes ", clausulas_analis_prev)
            print("LISTA DE ALL PREVIO ", clausulas_analis_prev)

            # TODO: CUANDO SE ELIMINA AUN NO FUNCIONA; FALTA QUE SEPA CUAND
            # SE ESTA ELIMINANDO el registro

            # Al inicio no hay requisitos seleccionados, por ende pasa directo
            if(len(clausulas_analis_prev)):
                for item in ids_clause_list_all:
                    # print ("ENTRO ", item)
                    if str(item) not in clausulas_analis_prev:
                        # print ("NO ESTA ", item)
                        ids_clause_list.append(item)
            else:
                ids_clause_list = ids_clause_list_all
            ids_clause_list = ids_clause_list_all

            self.diagnostic4_ids = self._default_diagnostic_line_ids_v2(
                '4.1', ids_clause_list)
            self.diagnostic4_1_ids = self._default_diagnostic_line_ids_v2(
                '4.2', ids_clause_list)
            self.diagnostic4_2_ids = self._default_diagnostic_line_ids_v2(
                '4.3', ids_clause_list)
            self.diagnostic4_3_1_ids = self._default_diagnostic_line_ids_v2(
                '4.4.1', ids_clause_list)
            self.diagnostic4_3_2_ids = self._default_diagnostic_line_ids_v2(
                '4.4.2', ids_clause_list)
            # print("2 ---->",self.diagnostic4_ids.ids)
            self.diagnostic5_1_ids = self._default_diagnostic_line_ids_v2(
                '5.1.1', ids_clause_list)
            self.diagnostic5_2_ids = self._default_diagnostic_line_ids_v2(
                '5.1.2', ids_clause_list)
            self.diagnostic5_1_1_ids = self._default_diagnostic_line_ids_v2(
                '5.2.1', ids_clause_list)
            self.diagnostic5_1_2_ids = self._default_diagnostic_line_ids_v2(
                '5.2.2', ids_clause_list)
            self.diagnostic5_2_1_ids = self._default_diagnostic_line_ids_v2(
                '5.3', ids_clause_list)

            self.diagnostic6_1_ids = self._default_diagnostic_line_ids_v2(
                '6.1.1', ids_clause_list)
            self.diagnostic6_2_ids = self._default_diagnostic_line_ids_v2(
                '6.1.2', ids_clause_list)
            self.diagnostic6_1_1_ids = self._default_diagnostic_line_ids_v2(
                '6.2.1', ids_clause_list)
            self.diagnostic6_1_2_ids = self._default_diagnostic_line_ids_v2(
                '6.2.2', ids_clause_list)
            self.diagnostic6_2_1_ids = self._default_diagnostic_line_ids_v2(
                '6.3', ids_clause_list)

            self.diagnostic7_1_1_ids = self._default_diagnostic_line_ids_v2(
                '7.1.1', ids_clause_list)
            self.diagnostic7_1_2_ids = self._default_diagnostic_line_ids_v2(
                '7.1.2', ids_clause_list)
            self.diagnostic7_1_3_ids = self._default_diagnostic_line_ids_v2(
                '7.1.3', ids_clause_list)
            self.diagnostic7_1_4_ids = self._default_diagnostic_line_ids_v2(
                '7.1.4', ids_clause_list)
            self.diagnostic7_1_5_1_ids = self._default_diagnostic_line_ids_v2(
                '7.1.5.1', ids_clause_list)
            self.diagnostic7_1_5_2_ids = self._default_diagnostic_line_ids_v2(
                '7.1.5.2', ids_clause_list)
            self.diagnostic7_1_6_ids = self._default_diagnostic_line_ids_v2(
                '7.1.6', ids_clause_list)
            self.diagnostic7_2_1_ids = self._default_diagnostic_line_ids_v2(
                '7.2', ids_clause_list)
            self.diagnostic7_3_1_ids = self._default_diagnostic_line_ids_v2(
                '7.3', ids_clause_list)
            self.diagnostic7_4_1_ids = self._default_diagnostic_line_ids_v2(
                '7.4', ids_clause_list)
            self.diagnostic7_5_1_ids = self._default_diagnostic_line_ids_v2(
                '7.5.1', ids_clause_list)
            self.diagnostic7_5_2_ids = self._default_diagnostic_line_ids_v2(
                '7.5.2', ids_clause_list)
            self.diagnostic7_5_3_1_ids = self._default_diagnostic_line_ids_v2(
                '7.5.3.1', ids_clause_list)
            self.diagnostic7_5_3_2_ids = self._default_diagnostic_line_ids_v2(
                '7.5.3.2', ids_clause_list)

            self.diagnostic8_1_1_ids = self._default_diagnostic_line_ids_v2(
                '8.1', ids_clause_list)
            self.diagnostic8_2_1_ids = self._default_diagnostic_line_ids_v2(
                '8.2.1', ids_clause_list)
            self.diagnostic8_2_2_ids = self._default_diagnostic_line_ids_v2(
                '8.2.2', ids_clause_list)
            self.diagnostic8_2_2_3_1_ids = self._default_diagnostic_line_ids_v2(
                '8.2.3.1', ids_clause_list)
            self.diagnostic8_2_2_3_2_ids = self._default_diagnostic_line_ids_v2(
                '8.2.3.2', ids_clause_list)
            self.diagnostic8_2_4_ids = self._default_diagnostic_line_ids_v2(
                '8.2.4', ids_clause_list)
            self.diagnostic8_3_1_ids = self._default_diagnostic_line_ids_v2(
                '8.3.1', ids_clause_list)
            self.diagnostic8_3_2_ids = self._default_diagnostic_line_ids_v2(
                '8.3.2', ids_clause_list)
            self.diagnostic8_3_3_ids = self._default_diagnostic_line_ids_v2(
                '8.3.3', ids_clause_list)
            self.diagnostic8_3_4_ids = self._default_diagnostic_line_ids_v2(
                '8.3.4', ids_clause_list)
            self.diagnostic8_3_5_ids = self._default_diagnostic_line_ids_v2(
                '8.3.5', ids_clause_list)
            self.diagnostic8_3_6_ids = self._default_diagnostic_line_ids_v2(
                '8.3.6', ids_clause_list)
            self.diagnostic8_4_1_ids = self._default_diagnostic_line_ids_v2(
                '8.4.1', ids_clause_list)
            self.diagnostic8_4_2_ids = self._default_diagnostic_line_ids_v2(
                '8.4.2', ids_clause_list)
            self.diagnostic8_4_3_ids = self._default_diagnostic_line_ids_v2(
                '8.4.3', ids_clause_list)
            self.diagnostic8_5_1_ids = self._default_diagnostic_line_ids_v2(
                '8.5.1', ids_clause_list)
            self.diagnostic8_5_2_ids = self._default_diagnostic_line_ids_v2(
                '8.5.2', ids_clause_list)
            self.diagnostic8_5_3_ids = self._default_diagnostic_line_ids_v2(
                '8.5.3', ids_clause_list)
            self.diagnostic8_5_4_ids = self._default_diagnostic_line_ids_v2(
                '8.5.4', ids_clause_list)
            self.diagnostic8_5_5_ids = self._default_diagnostic_line_ids_v2(
                '8.5.5', ids_clause_list)
            self.diagnostic8_5_6_ids = self._default_diagnostic_line_ids_v2(
                '8.5.6', ids_clause_list)
            self.diagnostic8_6_1_ids = self._default_diagnostic_line_ids_v2(
                '8.6.1', ids_clause_list)
            self.diagnostic8_7_1_ids = self._default_diagnostic_line_ids_v2(
                '8.7.1', ids_clause_list)
            self.diagnostic8_7_2_ids = self._default_diagnostic_line_ids_v2(
                '8.7.2', ids_clause_list)

            self.diagnostic9_1_1_ids = self._default_diagnostic_line_ids_v2(
                '9.1.1', ids_clause_list)
            self.diagnostic9_1_2_ids = self._default_diagnostic_line_ids_v2(
                '9.1.2', ids_clause_list)
            self.diagnostic9_1_3_ids = self._default_diagnostic_line_ids_v2(
                '9.1.3', ids_clause_list)
            self.diagnostic9_2_1_ids = self._default_diagnostic_line_ids_v2(
                '9.2.1', ids_clause_list)
            self.diagnostic9_2_2_ids = self._default_diagnostic_line_ids_v2(
                '9.2.2', ids_clause_list)
            self.diagnostic9_3_1_ids = self._default_diagnostic_line_ids_v2(
                '9.3.1', ids_clause_list)
            self.diagnostic9_3_2_ids = self._default_diagnostic_line_ids_v2(
                '9.3.2', ids_clause_list)
            self.diagnostic9_3_3_ids = self._default_diagnostic_line_ids_v2(
                '9.3.3', ids_clause_list)

            self.diagnostic10_1_1_ids = self._default_diagnostic_line_ids_v2(
                '10.1', ids_clause_list)
            self.diagnostic10_2_1_ids = self._default_diagnostic_line_ids_v2(
                '10.2.1', ids_clause_list)
            self.diagnostic10_2_2_ids = self._default_diagnostic_line_ids_v2(
                '10.2.2', ids_clause_list)
            self.diagnostic10_3_1_ids = self._default_diagnostic_line_ids_v2(
                '10.3', ids_clause_list)
        else:
            self.diagnostic4_ids = False
            self.diagnostic4_1_ids = False
            self.diagnostic4_2_ids = False
            self.diagnostic4_3_1_ids = False
            self.diagnostic4_3_2_ids = False

            self.diagnostic5_1_ids = False
            self.diagnostic5_2_ids = False
            self.diagnostic5_1_1_ids = False
            self.diagnostic5_1_2_ids = False
            self.diagnostic5_2_ids = False

            self.diagnostic6_1_ids = False
            self.diagnostic6_2_ids = False
            self.diagnostic6_1_1_ids = False
            self.diagnostic6_1_2_ids = False
            self.diagnostic6_2_ids = False

            self.diagnostic7_1_1_ids = False
            self.diagnostic7_1_2_ids = False
            self.diagnostic7_1_3_ids = False
            self.diagnostic7_1_4_ids = False
            self.diagnostic7_1_5_1_ids = False
            self.diagnostic7_1_5_2_ids = False
            self.diagnostic7_1_6_ids = False
            self.diagnostic7_2_1_ids = False
            self.diagnostic7_3_1_ids = False
            self.diagnostic7_4_1_ids = False
            self.diagnostic7_5_1_ids = False
            self.diagnostic7_5_2_ids = False
            self.diagnostic7_5_3_1_ids = False
            self.diagnostic7_5_3_2_ids = False

            self.diagnostic8_1_1_ids = False
            self.diagnostic8_2_1_ids = False
            self.diagnostic8_2_2_ids = False
            self.diagnostic8_2_2_3_1_ids = False
            self.diagnostic8_2_2_3_2_ids = False
            self.diagnostic8_2_4_ids = False
            self.diagnostic8_3_1_ids = False
            self.diagnostic8_3_2_ids = False
            self.diagnostic8_3_3_ids = False
            self.diagnostic8_3_4_ids = False
            self.diagnostic8_3_5_ids = False
            self.diagnostic8_3_6_ids = False
            self.diagnostic8_4_1_ids = False
            self.diagnostic8_4_2_ids = False
            self.diagnostic8_4_3_ids = False
            self.diagnostic8_5_1_ids = False
            self.diagnostic8_5_2_ids = False
            self.diagnostic8_5_3_ids = False
            self.diagnostic8_5_4_ids = False
            self.diagnostic8_5_5_ids = False
            self.diagnostic8_5_6_ids = False
            self.diagnostic8_6_1_ids = False
            self.diagnostic8_7_1_ids = False
            self.diagnostic8_7_2_ids = False

            self.diagnostic9_1_1_ids = False
            self.diagnostic9_1_2_ids = False
            self.diagnostic9_1_3_ids = False
            self.diagnostic9_2_1_ids = False
            self.diagnostic9_2_2_ids = False
            self.diagnostic9_3_1_ids = False
            self.diagnostic9_3_2_ids = False
            self.diagnostic9_3_3_ids = False

            # self.diagnostic10_ids = False
            # self.diagnostic10_1_ids = False
            # self.diagnostic10_2_ids = False

        # Cambia de estado al analizar detalladamente
        self.state = 'evaluate'

    def change_state_draft(self):
        """ Cambia el estado de evaluate a draft,
        cuando son seleccionadas las clausulas
        """
        self.state = 'draft'

    def change_state_validate(self):
        """ Cambia el estado de evaluate a validate.
        Esto culmina todo el proceso
        """
        self.state = 'validate'
